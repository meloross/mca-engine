from __future__ import annotations

import csv
from collections import Counter
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore[import-untyped]

from app.exports.schemas import ExportFilters, ExportType

COMPLIANCE_DISCLAIMER = (
    "This export contains public-record legal market intelligence and/or consented opt-in leads. "
    "It is not a cold-call list. Users are responsible for complying with attorney advertising, "
    "solicitation, privacy, TCPA, CAN-SPAM, and other applicable rules."
)


def rows_to_csv(headers: list[str], rows: list[dict[str, object]]) -> bytes:
    stream = StringIO()
    writer = csv.DictWriter(stream, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({header: _cell_value(row.get(header)) for header in headers})
    return stream.getvalue().encode("utf-8-sig")


def rows_to_xlsx(
    *,
    headers: list[str],
    rows: list[dict[str, object]],
    export_type: ExportType,
    export_timestamp: datetime,
    filters: ExportFilters,
    omitted_counts: dict[str, int],
) -> bytes:
    workbook = Workbook()
    signals_sheet = workbook.active
    signals_sheet.title = "Signals" if export_type == "signals" else "Form Leads"
    _write_data_sheet(signals_sheet, headers, rows)
    _write_summary_sheet(
        workbook=workbook,
        rows=rows,
        export_timestamp=export_timestamp,
        filters=filters,
        omitted_counts=omitted_counts,
    )
    _write_metadata_sheet(
        workbook=workbook,
        export_type=export_type,
        export_timestamp=export_timestamp,
    )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_data_sheet(
    worksheet: Worksheet,
    headers: list[str],
    rows: list[dict[str, object]],
) -> None:
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        worksheet.append([_cell_value(row.get(header)) for header in headers])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _autosize_columns(worksheet)


def _write_summary_sheet(
    *,
    workbook: Workbook,
    rows: list[dict[str, object]],
    export_timestamp: datetime,
    filters: ExportFilters,
    omitted_counts: dict[str, int],
) -> None:
    worksheet = workbook.create_sheet("Summary")
    grade_counts = Counter(str(row.get("grade", "")) for row in rows)
    summary_rows: list[tuple[str, object]] = [
        ("total rows", len(rows)),
        ("export timestamp", export_timestamp.isoformat()),
        ("filters used", _filters_label(filters)),
        ("total A_PLUS", grade_counts.get("A_PLUS", 0)),
        ("total A", grade_counts.get("A", 0)),
        ("total B", grade_counts.get("B", 0)),
        ("total C", grade_counts.get("C", 0)),
        ("total D", grade_counts.get("D", 0)),
        ("excluded count omitted", omitted_counts.get("excluded", 0)),
        ("suppressed count omitted", omitted_counts.get("suppressed", 0)),
    ]
    for row in summary_rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    _autosize_columns(worksheet)


def _write_metadata_sheet(
    *,
    workbook: Workbook,
    export_type: ExportType,
    export_timestamp: datetime,
) -> None:
    worksheet = workbook.create_sheet("Export Metadata")
    rows = [
        ("app name", "MCA Legal Signal Engine"),
        ("export type", export_type),
        ("export timestamp", export_timestamp.isoformat()),
        ("source note", "Generated from local database records and configured export filters."),
        ("compliance disclaimer", COMPLIANCE_DISCLAIMER),
    ]
    for row in rows:
        worksheet.append(row)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    _autosize_columns(worksheet)


def _filters_label(filters: ExportFilters) -> str:
    parts = [
        f"{key}={value}"
        for key, value in filters.as_metadata().items()
        if value not in ("", False)
    ]
    return "; ".join(parts) if parts else "none"


def _autosize_columns(worksheet: Worksheet) -> None:
    for column_cells in worksheet.columns:
        first_cell = column_cells[0]
        column_letter = first_cell.column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list | tuple | set):
        return "; ".join(str(item) for item in value)
    return str(value)
