from __future__ import annotations

from collections.abc import Iterator, Sequence
from contextlib import AbstractContextManager
from datetime import UTC, date, datetime
from pathlib import Path
from types import TracebackType
from typing import cast
from uuid import UUID, uuid4

from fastapi.testclient import TestClient
from openpyxl import load_workbook  # type: ignore[import-untyped]
from pytest import CaptureFixture, MonkeyPatch
from sqlalchemy.orm import Session

from app.api import exports as exports_api
from app.db import get_session
from app.exports import ExportFilters, ExportResult, export_signals_bytes
from app.exports.lead_exporter import PUBLIC_SIGNAL_COLUMNS, load_public_signal_export_rows
from app.main import app
from app.models import (
    Case,
    CaseDocument,
    LeadSignal,
    LeadSignalGrade,
    LeadSignalStatus,
    SignalType,
    Source,
    SourceType,
)
from app.models.enums import AccessMethod
from app.scripts.export_leads import run


def test_csv_export_creates_expected_headers() -> None:
    source, case, signal = _signal_fixture()
    session = _FakeSession(
        main=[signal],
        objects={Source: {source.id: source}, Case: {case.id: case}},
    )

    result = export_signals_bytes(
        cast(Session, session),
        filters=ExportFilters(),
        export_format="csv",
    )
    header = result.content.decode("utf-8-sig").splitlines()[0].split(",")

    assert header == PUBLIC_SIGNAL_COLUMNS
    assert result.row_count == 1


def test_xlsx_export_creates_expected_sheets(tmp_path: Path) -> None:
    source, case, signal = _signal_fixture()
    session = _FakeSession(
        main=[signal],
        objects={Source: {source.id: source}, Case: {case.id: case}},
    )

    result = export_signals_bytes(
        cast(Session, session),
        filters=ExportFilters(),
        export_format="xlsx",
    )
    output = tmp_path / "signals.xlsx"
    output.write_bytes(result.content)
    workbook = load_workbook(output)

    assert workbook.sheetnames == ["Signals", "Summary", "Export Metadata"]
    assert workbook["Signals"].freeze_panes == "A2"
    assert workbook["Signals"].auto_filter.ref is not None


def test_high_value_state_and_min_score_filters() -> None:
    ny_source, ny_case, ny_signal = _signal_fixture(score=95, grade=LeadSignalGrade.A_PLUS)
    _, _, low_signal = _signal_fixture(
        state="NY",
        score=40,
        grade=LeadSignalGrade.C,
        business_name="Low Score Merchant LLC",
    )
    _, _, fl_signal = _signal_fixture(
        state="FL",
        score=95,
        grade=LeadSignalGrade.A_PLUS,
        business_name="Florida Merchant LLC",
    )
    session = _FakeSession(
        main=[ny_signal, low_signal, fl_signal],
        objects={Source: {ny_source.id: ny_source}, Case: {ny_case.id: ny_case}},
    )

    rows, _ = load_public_signal_export_rows(
        cast(Session, session),
        ExportFilters.from_state(state="NY", only_high_value=True, min_score=75),
    )

    assert [row["business_name"] for row in rows] == ["Demo Merchant LLC"]


def test_suppressed_and_excluded_are_omitted_by_default() -> None:
    source, case, new_signal = _signal_fixture()
    _, _, suppressed = _signal_fixture(
        business_name="Suppressed Merchant LLC",
        status=LeadSignalStatus.SUPPRESSED,
    )
    _, _, excluded = _signal_fixture(
        business_name="Excluded Merchant LLC",
        status=LeadSignalStatus.EXCLUDED,
        grade=LeadSignalGrade.EXCLUDE,
        score=0,
    )
    session = _FakeSession(
        main=[new_signal, suppressed, excluded],
        objects={Source: {source.id: source}, Case: {case.id: case}},
    )

    rows, omitted = load_public_signal_export_rows(cast(Session, session), ExportFilters())

    assert [row["business_name"] for row in rows] == ["Demo Merchant LLC"]
    assert omitted == {"excluded": 1, "suppressed": 1}


def test_cli_export_command_writes_file(
    tmp_path: Path,
    capsys: CaptureFixture[str],
) -> None:
    source, case, signal = _signal_fixture()
    session = _FakeSession(
        main=[signal],
        objects={Source: {source.id: source}, Case: {case.id: case}},
    )
    output = tmp_path / "ny_high_value_mca_signals.csv"

    exit_code = run(
        [
            "--type",
            "signals",
            "--format",
            "csv",
            "--state",
            "NY",
            "--only-high-value",
            "--output",
            str(output),
        ],
        session_factory=lambda: _SessionContext(session),
    )

    captured = capsys.readouterr()
    assert exit_code == 0
    assert output.exists()
    assert "Exported 1 rows" in captured.out


def test_api_file_response_content_types(monkeypatch: MonkeyPatch) -> None:
    timestamp = datetime(2026, 6, 10, 14, 30, tzinfo=UTC)
    fake_csv = ExportResult(
        content=b"signal_id\n",
        filename="mca_signals_NY_A_PLUS_A_2026-06-10_143000.csv",
        media_type="text/csv",
        row_count=1,
        export_timestamp=timestamp,
        filters=ExportFilters.from_state(state="NY", only_high_value=True),
    )
    fake_xlsx = ExportResult(
        content=b"PK\x03\x04",
        filename="mca_form_leads_ALL_A_PLUS_A_2026-06-10_143000.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        row_count=1,
        export_timestamp=timestamp,
        filters=ExportFilters(only_high_value=True),
    )
    monkeypatch.setattr(exports_api, "export_signals_bytes", lambda *args, **kwargs: fake_csv)
    monkeypatch.setattr(exports_api, "export_form_leads_bytes", lambda *args, **kwargs: fake_xlsx)
    app.dependency_overrides[get_session] = lambda: object()
    client = TestClient(app)

    try:
        csv_response = client.get("/exports/signals.csv?state=NY&only_high_value=true")
        xlsx_response = client.get("/exports/form-leads.xlsx?only_high_value=true")
    finally:
        app.dependency_overrides.clear()

    assert csv_response.headers["content-type"].startswith("text/csv")
    assert xlsx_response.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in csv_response.headers["content-disposition"]


def _signal_fixture(
    *,
    state: str = "NY",
    score: int = 95,
    grade: LeadSignalGrade = LeadSignalGrade.A_PLUS,
    status: LeadSignalStatus = LeadSignalStatus.NEW,
    business_name: str = "Demo Merchant LLC",
) -> tuple[Source, Case, LeadSignal]:
    source_id = uuid4()
    case_id = uuid4()
    signal_id = uuid4()
    source = Source(
        id=source_id,
        name=f"{state} Demo Source",
        state=state,
        source_type=SourceType.COURT_NEW_CASES,
        base_url="https://demo.local",
        access_method=AccessMethod.MOCK,
        requires_login=False,
        requires_payment=False,
    )
    case = Case(
        id=case_id,
        state=state,
        county="Kings" if state == "NY" else "Miami-Dade",
        court_name="Demo Commercial Court",
        case_number=f"{state}-2026-001",
        case_type="Commercial",
        filing_date=date(2026, 6, 10),
        last_activity_date=date(2026, 6, 10),
        caption=f"Cloudfund LLC v. {business_name}",
        plaintiff_names=["Cloudfund LLC"],
        defendant_names=[business_name],
        attorney_names=[],
        source_id=source_id,
        source_url="https://demo.local/case",
        normalized_key=f"{state}:CASE:001",
    )
    signal = LeadSignal(
        id=signal_id,
        signal_type=SignalType.LITIGATION_NEW_CASE,
        state=state,
        county=case.county,
        business_name=business_name,
        normalized_business_name="DEMO MERCHANT",
        funder_name="Cloudfund",
        case_id=case_id,
        signal_date=date(2026, 6, 10),
        title=f"Demo signal for {business_name}",
        summary="Merchant cash advance and daily ACH dispute.",
        score=score,
        risk_score=15,
        grade=grade,
        status=status,
        compliance_flags=["mock"],
        source_id=source_id,
        source_url="https://demo.local/signal",
        created_at=datetime(2026, 6, 10, tzinfo=UTC),
        updated_at=datetime(2026, 6, 10, tzinfo=UTC),
    )
    return source, case, signal


class _FakeScalarResult:
    def __init__(self, values: Sequence[object]) -> None:
        self._values = list(values)

    def all(self) -> list[object]:
        return self._values

    def __iter__(self) -> Iterator[object]:
        return iter(self._values)


class _FakeSession:
    def __init__(
        self,
        *,
        main: list[object],
        objects: dict[type[object], dict[UUID, object]] | None = None,
        documents: list[CaseDocument] | None = None,
    ) -> None:
        self._main = main
        self._objects = objects or {}
        self._documents = documents or []
        self._scalar_calls = 0

    def scalars(self, statement: object) -> _FakeScalarResult:
        self._scalar_calls += 1
        if self._scalar_calls == 1:
            return _FakeScalarResult(self._main)
        return _FakeScalarResult(self._documents)

    def get(self, model: type[object], object_id: UUID | None) -> object | None:
        if object_id is None:
            return None
        return self._objects.get(model, {}).get(object_id)

    def scalar(self, statement: object) -> object | None:
        return None


class _SessionContext(AbstractContextManager[Session]):
    def __init__(self, session: _FakeSession) -> None:
        self._session = session

    def __enter__(self) -> Session:
        return cast(Session, self._session)

    def __exit__(
        self,
        exc_type: type[BaseException] | None,
        exc_value: BaseException | None,
        traceback: TracebackType | None,
    ) -> None:
        return None
